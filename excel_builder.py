# -*- coding: utf-8 -*-
"""Bouwt het Excelbestand: tabblad per shop, gesorteerd op score."""

import logging

from openpyxl import Workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

import config
import scoring

log = logging.getLogger("bierscraper")

# (veldnaam intern, kolomkop, breedte)
BASE_COLUMNS = [
    ("brouwerij", "Brouwerij", 24),
    ("naam", "Naam", 38),
    ("inhoud_cl", "Inhoud (cl)", 11),
    ("land", "Land", 14),
    ("abv", "ABV %", 8),
    ("stijl", "Bierstijl", 34),
    ("untappd", "Untappd score", 13),
    ("untappd_aantal", "Untappd aantal", 13),
    ("prijs", "Prijs (€)", 10),
    ("weblink", "Weblink", 45),
]

HEADER_FILL = PatternFill("solid", start_color="1F4E44")
HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF")
BODY_FONT = Font(name="Arial")
STRONG_FILL = PatternFill("solid", start_color="FFF2CC")  # sterke voorkeur licht geel
# Let op: voor VOORWAARDELIJKE opmaak (dxf) moet zowel start- als eindkleur
# gezet zijn, anders toont Excel geen vulling (bekende openpyxl-eigenaardigheid)
RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
GREEN_FILL = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")  # felgroen


def build_workbook(all_beers, sites):
    scoring.compute_scores(all_beers)
    price_lookup = scoring.build_price_lookup(all_beers)

    wb = Workbook()
    wb.remove(wb.active)

    for site in sites:
        beers = sorted(
            all_beers.get(site["key"], []),
            key=lambda b: b.get("score") or 0,
            reverse=True,
        )
        _build_sheet(wb, site, beers, sites, price_lookup)

    return wb


def _build_sheet(wb, site, beers, sites, price_lookup):
    ws = wb.create_sheet(site["label"][:31])

    # Alleen kolommen opnemen waarvoor deze shop daadwerkelijk data heeft;
    # bij elke run opnieuw bepaald, dus als de shop het veld later toevoegt
    # verschijnt de kolom vanzelf.
    columns = [(f, h, w) for f, h, w in BASE_COLUMNS
               if any(b.get(f) is not None for b in beers)]

    other_sites = [s for s in sites if s["key"] != site["key"]]

    headers = [h for _, h, _ in columns] + ["Score (0-100)"] \
        + [f"Prijs {s['label']} (€)" for s in other_sites]

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill, cell.font = HEADER_FILL, HEADER_FONT
        cell.alignment = Alignment(vertical="center", wrap_text=True)

    price_col_idx = next((i + 1 for i, (f, _, _) in enumerate(columns) if f == "prijs"), None)
    score_col_idx = len(columns) + 1

    for row_idx, beer in enumerate(beers, start=2):
        for col_idx, (field, _, _) in enumerate(columns, start=1):
            value = beer.get(field)
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = BODY_FONT
            if field == "weblink" and value:
                cell.hyperlink = value
                cell.font = Font(name="Arial", color="0563C1", underline="single")
            if field == "naam" and beer.get("sterke_voorkeur"):
                cell.fill = STRONG_FILL
            if field in ("prijs",):
                cell.number_format = "€ #,##0.00"
            if field == "untappd" and value is not None:
                cell.number_format = "0.00"

        ws.cell(row=row_idx, column=score_col_idx, value=beer.get("score")).font = \
            Font(name="Arial", bold=True)

        for offset, other in enumerate(other_sites):
            price = scoring.find_price(beer, price_lookup.get(other["key"], {}))
            cell = ws.cell(row=row_idx, column=score_col_idx + 1 + offset, value=price)
            cell.font = BODY_FONT
            cell.number_format = "€ #,##0.00"

    last_row = max(len(beers) + 1, 2)

    # --- voorwaardelijke opmaak op de prijskolommen van de andere shops ---
    if price_col_idx and beers:
        own = f"${get_column_letter(price_col_idx)}2"
        for offset in range(len(other_sites)):
            col_letter = get_column_letter(score_col_idx + 1 + offset)
            rng = f"{col_letter}2:{col_letter}{last_row}"
            other_cell = f"{col_letter}2"
            ws.conditional_formatting.add(rng, FormulaRule(
                formula=[f'AND({other_cell}<>"",{own}<>"",{other_cell}>{own})'],
                fill=RED_FILL, stopIfTrue=False))
            ws.conditional_formatting.add(rng, FormulaRule(
                formula=[f'AND({other_cell}<>"",{own}<>"",{other_cell}<{own})'],
                fill=GREEN_FILL, stopIfTrue=False))

    # --- opmaak ---
    widths = [w for _, _, w in columns] + [13] + [16] * len(other_sites)
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{last_row}"

    log.info("Tabblad '%s': %d bieren, %d kolommen", site["label"], len(beers), len(headers))
